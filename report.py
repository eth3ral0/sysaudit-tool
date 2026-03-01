import os
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class ReportGenerator:
    """Genere les rapports (PDF, Excel)"""

    def __init__(self, data):
        self.data = data
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_dir = "reports"
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_pdf(self, filename=None):
        """Genere un PDF du rapport"""
        if not filename:
            filename = f"audit_it_{self.timestamp}.pdf"
        filepath = os.path.join(self.output_dir, filename)

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Titre
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1F2121"),
            spaceAfter=30,
        )
        elements.append(Paragraph("Audit Informatique", title_style))
        elements.append(Spacer(1, 0.3 * inch))

        # Infos de base
        basic = self.data.get("basic", {})
        elements.append(Paragraph(f"<b>Poste :</b> {basic.get('hostname')}", styles["Normal"]))
        elements.append(Paragraph(f"<b>OS :</b> {basic.get('os')} {basic.get('os_version')}", styles["Normal"]))
        elements.append(Paragraph(f"<b>Date du scan :</b> {basic.get('scan_date')}", styles["Normal"]))
        elements.append(Spacer(1, 0.2 * inch))

        # CPU
        cpu = self.data.get("cpu", {})
        elements.append(Paragraph("<b>Processeur</b>", styles["Heading2"]))
        elements.append(Paragraph(f"Cores physiques: {cpu.get('cpu_count_physical')}", styles["Normal"]))
        elements.append(Paragraph(f"Cores logiques: {cpu.get('cpu_count_logical')}", styles["Normal"]))
        elements.append(Paragraph(f"Frequence: {cpu.get('cpu_freq_mhz')} MHz", styles["Normal"]))
        elements.append(Paragraph(f"Utilisation: {cpu.get('cpu_percent')}%", styles["Normal"]))
        elements.append(Spacer(1, 0.2 * inch))

        # Memoire
        mem = self.data.get("memory", {})
        elements.append(Paragraph("<b>Memoire</b>", styles["Heading2"]))
        elements.append(Paragraph(f"Total: {mem.get('memory_total_gb')} GB", styles["Normal"]))
        elements.append(Paragraph(f"Utilisee: {mem.get('memory_used_gb')} GB ({mem.get('memory_percent')}%)", styles["Normal"]))
        elements.append(Spacer(1, 0.2 * inch))

        # Disques
        disks = self.data.get("disk", [])
        if disks:
            elements.append(Paragraph("<b>Disques</b>", styles["Heading2"]))
            disk_data = [["Disque", "Total (GB)", "Utilise (GB)", "% Utilise"]]
            for disk in disks:
                disk_data.append([
                    disk["device"],
                    str(disk["total_gb"]),
                    str(disk["used_gb"]),
                    str(disk["percent"]) + "%",
                ])
            table = Table(disk_data)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)

                # Hotes reseau (si disponible)
        network_hosts = self.data.get("network_hosts", [])
        if network_hosts:
            elements.append(Spacer(1, 0.2 * inch))
            elements.append(Paragraph("<b>Hotes reseau detectes</b>", styles["Heading2"]))
            for host in network_hosts:
                ip = host.get('ip', 'N/A')
                status = host.get('status', 'N/A')
                elements.append(Paragraph(f"- {ip} ({status})", styles["Normal"]))


        summary = self.data.get("health_summary", "")
        if summary:
            elements.append(Spacer(1, 0.2 * inch))
            elements.append(Paragraph("<b>Synthese de l'etat du poste</b>", styles["Heading2"]))
            elements.append(Paragraph(summary, styles["Normal"]))


        doc.build(elements)
        print(f"PDF genere: {filepath}")
        return filepath

    def generate_excel(self, filename=None):
        """Genere un Excel du rapport"""
        if not filename:
            filename = f"audit_it_{self.timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit IT"


        ws["A1"] = "AUDIT INFORMATIQUE"
        ws["A1"].font = Font(bold=True, size=14)

        # Infos de base
        basic = self.data.get("basic", {})
        row = 3
        ws[f"A{row}"] = "Poste"
        ws[f"B{row}"] = basic.get("hostname")
        row += 1
        ws[f"A{row}"] = "OS"
        ws[f"B{row}"] = f"{basic.get('os')} {basic.get('os_version')}"
        row += 1
        ws[f"A{row}"] = "Date du scan"
        ws[f"B{row}"] = basic.get("scan_date")
        row += 2

        # CPU
        cpu = self.data.get("cpu", {})
        ws[f"A{row}"] = "PROCESSEUR"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        ws[f"A{row}"] = "Cores physiques"
        ws[f"B{row}"] = cpu.get("cpu_count_physical")
        row += 1
        ws[f"A{row}"] = "Cores logiques"
        ws[f"B{row}"] = cpu.get("cpu_count_logical")
        row += 2

        # Memoire
        mem = self.data.get("memory", {})
        ws[f"A{row}"] = "MEMOIRE"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        ws[f"A{row}"] = "Total (GB)"
        ws[f"B{row}"] = mem.get("memory_total_gb")
        row += 1
        ws[f"A{row}"] = "Utilisee (GB)"
        ws[f"B{row}"] = mem.get("memory_used_gb")
        row += 1
        ws[f"A{row}"] = "% Utilise"
        ws[f"B{row}"] = mem.get("memory_percent")

        # Auto-adjust columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25


            # Fonction helper pour colorier les cellules selon le %
        def get_color_fill(percent):
            """Retourne PatternFill rouge (>90%), orange (70-90%), vert (<70%)"""
            try:
                val = float(percent.rstrip('%'))
                if val > 90:
                    return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rouge
                elif val >= 70:
                    return PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
                else:
                    return PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Vert
            except:
                return PatternFill()

        # Section disques avec couleurs
        row += 2
        ws[f"A{row}"] = "DISQUES"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        # En-tete
        ws[f"A{row}"] = "Partition"
        ws[f"B{row}"] = "Total (GB)"
        ws[f"C{row}"] = "Utilise (GB)"
        ws[f"D{row}"] = "Libre (GB)"
        ws[f"E{row}"] = "% Utilise"
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row}"].font = Font(bold=True, color="FFFFFF")
            ws[f"{col}{row}"].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1

        # Donnees disques avec formatage des %
        for d in self.data.get("disk", []):
            ws[f"A{row}"] = d.get("device", "")
            ws[f"B{row}"] = f"{d.get('total_gb', 0):.2f}"
            ws[f"C{row}"] = f"{d.get('used_gb', 0):.2f}"
            ws[f"D{row}"] = f"{d.get('free_gb', 0):.2f}"

            # Cellule du pourcentage avec couleur
            percent_str = f"{d.get('percent', 0):.1f}%"
            ws[f"E{row}"] = percent_str
            ws[f"E{row}"].fill = get_color_fill(percent_str)
            ws[f"E{row}"].font = Font(bold=True, color="FFFFFF")  # Texte blanc pour plus de lisibilite


                    # Hotes reseau (si disponible)
        network_hosts = self.data.get("network_hosts", [])
        if network_hosts:
            row += 2
            ws[f"A{row}"] = "HOTES RESEAU DETECTES"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            for host in network_hosts:
                ip = host.get('ip', 'N/A')
                status = host.get('status', 'N/A')
                ws[f"A{row}"] = f"{ip}"
                ws[f"B{row}"] = f"{status}"
                row += 1

            row += 1
        # Sante du systeme (si disponible)
        row += 2
        ws[f"A{row}"] = "SYNTHESE ETAT DU POSTE"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        ws[f"A{row}"] = self.data.get("health_summary", "")

        wb.save(filepath)
        print(f"Excel genere: {filepath}")
        return filepath
